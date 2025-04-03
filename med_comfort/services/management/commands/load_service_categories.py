from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from services.models import ServiceCategory


class Command(BaseCommand):
    help = "Load service categories from an Excel file"

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename='data/service_categories.xlsx')
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, description = row
                category, created = ServiceCategory.objects.get_or_create(
                    name=name,
                    description=description
                )

                if created:
                    print(f"Added category: {name}")
                else:
                    print(f"Category already exists: {name}")

        except Exception as e:
            raise CommandError(f"Error while processing the file: {str(e)}")
