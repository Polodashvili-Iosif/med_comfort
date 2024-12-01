from datetime import timedelta

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from services.models import Service, ServiceCategory


class Command(BaseCommand):
    help = "Load services from an Excel file"

    def handle(self, *args, **options):
        try:
            wb = load_workbook(filename='data/services.xlsx')
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                category_name, name, description, duration = row
                hours, minutes, seconds = map(int, duration.split(':'))
                duration = timedelta(
                    hours=hours, minutes=minutes, seconds=seconds
                )
                category = ServiceCategory.objects.get(name=category_name)
                service, created = Service.objects.get_or_create(
                    name=name,
                    description=description,
                    duration=duration,
                    category=category
                )

                if created:
                    print(f"Added service: {name}")
                else:
                    print(f"Service already exists: {name}")

        except Exception as e:
            raise CommandError(f"Error while processing the file: {str(e)}")
